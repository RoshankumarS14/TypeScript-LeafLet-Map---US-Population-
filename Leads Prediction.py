import streamlit as st
import numpy as np
import pandas as pd
from plot_utils import plot_gauge_Balance, plot_gauge_APScale
from plotly.subplots import make_subplots
from openpyxl import load_workbook
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
from openpyxl.drawing.image import Image as XLImage
from PIL import Image 
import shutil
from plotly.io import to_image
import io
from openpyxl import Workbook

st.set_page_config(
    page_title="Leads Prediction",
    page_icon="📈",
    layout="centered",
    initial_sidebar_state="collapsed"
)

df = pd.read_excel("Silo-Data-(12-04-23).xlsx")
role_adjuster = pd.read_excel("Silo-Data-Classifications.xlsx")
states = df["ST"].unique()
silos = df["Silo"].unique()
ap_scale_silos = dict(df[["Silo","AP-Scale"]].values)

st.image("logo.png")

if 'company_name' not in st.session_state:
    st.session_state.company_name = ''
if 'user_role' not in st.session_state:
    st.session_state.user_role = ''

col_name,col_name_input,col_role,col_role_input = st.columns([0.5,1.6,0.25,1.7]) 
col_name.markdown(f"<div style='text-align: center; color: white; padding-top: 32px; font-size:18px;'>Company:</div>", unsafe_allow_html=True)
name_input_slot = col_name_input.empty()
st.session_state.company_name = name_input_slot.text_input('', '', key="Company_name")
col_role.markdown(f"<div style='text-align: center; color: white; padding-top: 32px; font-size:18px;'>Role:</div>", unsafe_allow_html=True)
role_input_slot = col_role_input.empty()
st.session_state.user_role = role_input_slot.text_input('', '', key="role")
if len(st.session_state.user_role) > 18:
   st.warning(f"Input is too long! Please limit your input to 18 characters.")

cols_campaign = st.columns([0.7, 0.1, 0.9] * 3) 
campaigns=["Full:","Half:","Quarter:"]
campaigns_values = []
    
for j in range(3):

    col_name = cols_campaign[j*3]
    col_dollar_sign = cols_campaign[j*3 + 1]
    col_input = cols_campaign[j*3 + 2]
    
    # Display the option
    col_name.markdown(f"<div style='text-align: center; color: white; padding-top: 32px; font-size:18px; margin-bottom:60px;'>{campaigns[j]}</div>", unsafe_allow_html=True)
    
    # Display the dollar sign
    col_dollar_sign.markdown(f"<div style='text-align: right; color: white; padding-top: 30px; font-size: 20px; margin-bottom:60px;'>$</div>", unsafe_allow_html=True)
    
    # Create the text input slot
    text_input_slot = col_input.empty()
    user_input = text_input_slot.text_input('', '', key=f'input_campaign{j}')
    campaigns_values.append(user_input)

role = st.selectbox("Select the job role:",role_adjuster["Role"],4)
input_states = st.multiselect("Select the states:",states)
input_silos = st.multiselect("Select the Silos:",silos)
input_budget = []

def calculate_rating(numbers):
    # Calculate the standard deviation
    std_dev = np.std(numbers)
    
    # Calculate the mean of the numbers
    mean = np.mean(numbers)
    
    # Normalize the standard deviation to a scale of 0 to 1
    normalized_std_dev = std_dev / mean if mean != 0 else 0
    
    # Calculate the rating
    rating = 100 * (1 - normalized_std_dev)
    
    # Make sure the rating is within the scale of 1 to 100
    rating = max(min(rating, 100), 1)
    
    return rating

st.text("Enter the budget for each of the media silos:")
# Calculate the number of rows
num_rows = len(input_silos) // 3
if len(input_silos) % 3 != 0:
    num_rows += 1

# For each row
for i in range(num_rows):
    # Create 9 columns: 3 for options, 3 for dollar signs, and 3 for inputs
    cols = st.columns([0.7, 0.1, 0.9] * 3)  # Repeat the pattern for 3 sets of columns
    
    # For each set of option, dollar sign, and input
    for j in range(3):
        # Calculate the index for the silo
        index = i * 3 + j
        
        # If the index is out of range, break the loop
        if index >= len(input_silos):
            break
        
        # Get the option
        option = input_silos[index]
        
        # Calculate the index for the columns
        col_option = cols[j*3]
        col_dollar = cols[j*3 + 1]
        col_input = cols[j*3 + 2]
        
        # Display the option
        col_option.markdown(f"<div style='text-align: center; color: white; padding-top: 32px; font-size:18px;'>{option}</div>", unsafe_allow_html=True)
        
        # Display the dollar sign
        col_dollar.markdown(f"<div style='text-align: right; color: white; padding-top: 30px; font-size: 20px;'>$</div>", unsafe_allow_html=True)
        
        # Create the text input slot
        text_input_slot = col_input.empty()
        user_input = text_input_slot.text_input('', '', key=f'input_{index}')
        input_budget.append(user_input)

def set_predict_leads():
    st.session_state.predict_leads = True

if 'predict_leads' not in st.session_state:
    st.session_state.predict_leads = False

st.button("Predict", on_click=set_predict_leads, use_container_width=True)

if st.session_state.predict_leads:
    leads = []
    average_CPLs = []
    input_budget = [float(i) for i in input_budget]
    for silo,budget in zip(input_silos,input_budget):
        CPLs = []
        for state in input_states:
            CPLs.append(df[(df["Silo"]==silo) & (df["ST"]==state)]["CPL"])
        CPLs = pd.Series([CPL.values[0] if len(CPL)>0 else None for CPL in CPLs]).dropna()
        if len(CPLs)==0:
            CPLs = pd.Series([df[(df["Silo"]==silo) & (df["ST"]=="US")]["CPL"]]).values[0]
        average_CPL = CPLs.sum()/len(CPLs)
        average_CPLs.append(average_CPL)
        leads.append(round(budget/average_CPL,1)) 
    AP_scales = [ap_scale_silos[silo]*lead for silo,lead in zip(input_silos,leads)]
    average_AP_scales = round(sum(AP_scales)/sum(leads),1)

    result = pd.DataFrame({"Silo":input_silos,"Budget":input_budget,"Average CPL":average_CPLs,"Leads":leads})
    result["Budget"] = result["Budget"].apply(lambda a: "$ " + '{:.2f}'.format(a))
    result["Average CPL"] = result["Average CPL"].apply(lambda a: "$ " + '{:.2f}'.format(a))
    result.index = np.arange(1,len(result)+1)
    # st.dataframe(result)
    # Convert the DataFrame to HTML and align all columns to the right
    df_html = result.to_html(classes='table table-striped')
    df_html = df_html.replace('<table ','<table style="text-align:right; margin-bottom:40px; margin-top:50px; width:95%;" ')

    col_df,col_gauge = st.columns([1.5,1])

    with col_df:
        # Display the DataFrame
        st.markdown(df_html, unsafe_allow_html=True)
        # st.write("AP Scale: "+str(average_AP_scales))
        # st.write("Total Budget: $"+str(sum(input_budget)))
        # st.write("Target Leads: "+str(round(sum(leads))*2))
        # st.write("Min Leads: "+str(round(sum(leads))))
    
        adjuster = role_adjuster[role_adjuster["Role"]==role]["Adjuster"].values[0]
        # Create formatted strings

        col_result1,col_result2 = st.columns([1,1])

        with col_result1:
            st.write(f"""
            <pre>
            <table style="border: none;">
            <tr><td style="text-align: left; border: none;">AP Scale:</td><td style="text-align: right; border: none;">{int(average_AP_scales*10)}</td></tr>
            <tr><td style="text-align: left; border: none;">Balance:</td><td style="text-align: right; border: none;">{int(calculate_rating(np.log([10000 if i>10000 else i for i in input_budget if i>50])))}</td></tr>
            <tr><td style="text-align: left; border: none;">Overhead Funds:</td><td style="text-align: right; border: none;">{round((1-(sum(input_budget)/float(campaigns_values[0])))*100)}</td></tr>
            </table>
            </pre>
            """, unsafe_allow_html=True)

        with col_result2:
            st.write(f"""
            <pre>
            <table style="border: none;">
            
            <tr><td style="text-align: left; border: none;">Target Leads:</td><td style="text-align: right; border: none;">{round(sum(leads)*adjuster)*2}</td></tr>
            <tr><td style="text-align: left; border: none;">Min Leads:</td><td style="text-align: right; border: none;">{round(sum(leads)*adjuster)}</td></tr>
            <tr><td style="text-align: left; border: none;">Min Budget:</td><td style="text-align: right; border: none;">${str(int(sum(input_budget)))}</td></tr>
            </table>
            </table>
            </pre>
            """, unsafe_allow_html=True)

    with col_gauge:
        trace1 = plot_gauge_APScale(int(average_AP_scales*10))
        trace2 = plot_gauge_Balance(int(calculate_rating(np.log([10000 if i>10000 else i for i in input_budget if i>50]))))
        # Create a subplot and add the gauges to it
        fig = make_subplots(rows=2, cols=1, specs=[[{'type': 'indicator'}],[{'type': 'indicator'}]],vertical_spacing=0.2)
        fig.append_trace(trace1, row=1, col=1)
        fig.append_trace(trace2, row=2, col=1)
        
        # Display the chart in Streamlit
        st.plotly_chart(fig, use_container_width=True)
        
    trace3 = plot_gauge_APScale(int(average_AP_scales*10),title="")
    fig2 = make_subplots(rows=1, cols=1, specs=[[{'type': 'indicator'}]])
    fig2.append_trace(trace3, row=1, col=1)
    
    # # Save the Plotly figure as an image file
    # fig2.write_image("fig.png")

    # # Open the image file
    # img = Image.open('fig.png')

    fig2_bytes = to_image(fig2, format="png")

    # Open the image file from memory
    img = Image.open(io.BytesIO(fig2_bytes))
    
    # Resize the image
    width, height = img.size
    new_width = 360
    new_height = 155
    img = img.resize((new_width, new_height))
    
    # # Save the resized image
    # img.save('resized_fig.png')

    # Save the resized image to a BytesIO object
    img_byte_arr = io.BytesIO()
    img.save(img_byte_arr, format='PNG')
    img_byte_arr = img_byte_arr.getvalue()
    
    # Load the workbook
    wb = load_workbook('New-Template.xlsx')
    
    # Select the sheet
    sheet = wb['juliabid'] 
    # del wb['juliabid']._images[:-2]
    # Create an Image object from BytesIO object
    img = XLImage(io.BytesIO(img_byte_arr))
    
    # Add the image to the sheet
    sheet.add_image(img, 'B36')
    
    # # Save the workbook
    # wb.save('New-Template.xlsx')

    # # Create a Pandas Excel writer using openpyxl as the engine
    # writer = pd.ExcelWriter(
    #     'New-Template.xlsx',
    #     engine='openpyxl',
    #     mode='a',
    #     if_sheet_exists='overlay',
    # )
     
    # # Write DataFrame to Excel from cell X11 for the first column
    # result.iloc[:, 0].to_excel(writer, sheet_name='juliabid', startrow=10, startcol=23, header=False, index=False)
    
    # # Write DataFrame to Excel from cell Y11 for the second column
    # pd.Series(input_budget).to_excel(writer, sheet_name='juliabid', startrow=10, startcol=24, header=False, index=False)
    
    # # Write DataFrame to Excel from cell AA11 for the third column
    # result.iloc[:, 3].to_excel(writer, sheet_name='juliabid', startrow=10, startcol=26, header=False, index=False)

    # # campaigns_values = [0 if i=="" else '{:.2f}'.format(float(i)) for i in campaigns_values] 
    # campaigns_values = [0 if i=="" else float(i) for i in campaigns_values] 
    # pd.Series(campaigns_values).to_excel(writer, sheet_name='juliabid', startrow=5, startcol=24, header=False, index=False)

     
    # pd.Series([st.session_state.company_name]).to_excel(writer, sheet_name='juliabid', startrow=4, startcol=24, header=False, index=False)
    # pd.Series([st.session_state.user_role]).to_excel(writer, sheet_name='juliabid', startrow=6, startcol=26, header=False, index=False)
    
    # # Save the workbook
    # writer.close()

    # Write DataFrame to Excel from cell X11 for the first column
    for i, value in enumerate(result.iloc[:, 0]):
        sheet.cell(row=i+11, column=24, value=value)
    
    # Write DataFrame to Excel from cell Y11 for the second column
    for i, value in enumerate(input_budget):
        sheet.cell(row=i+11, column=25, value=value)
    
    # Write DataFrame to Excel from cell AA11 for the third column
    for i, value in enumerate(result.iloc[:, 3]):
        sheet.cell(row=i+11, column=27, value=value)
        
    campaigns_values = [0 if i=="" else float(i) for i in campaigns_values] 
    # Write DataFrame to Excel from cell Y6 for the campaigns_values
    for i, value in enumerate(campaigns_values):
        sheet.cell(row=i+6, column=25, value=value if value != "" else 0.0)

    sheet.cell(row=5,column=25,value=st.session_state.company_name)
    sheet.cell(row=7,column=27,value=st.session_state.user_role)
    
    # Save the workbook to a BytesIO object
    excel_byte_arr = io.BytesIO()
    wb.save(excel_byte_arr)
        
    file_name = "TJD-" + st.session_state.company_name + ".xlsx"
    with open("New-Template.xlsx", "rb") as file:
         file_bytes = file.read()
    st.download_button(
        label="Create Campaign!",
        data=excel_byte_arr.getvalue(),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )


    






















